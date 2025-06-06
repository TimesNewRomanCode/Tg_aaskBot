import subprocess
import os, shutil
from openpyxl import load_workbook, Workbook
from copy import copy
import xlrd
from Database.db import GROUP_NAMES
from pdf2image import convert_from_path
import requests
from datetime import datetime, timedelta
from openpyxl.styles import Border, Side


async def download_and_generate_schedule():
    today = datetime.now()
    day_of_week = today.weekday()

    if day_of_week == 6:  # Воскресенье
        target_day = today + timedelta(days=1)
    else:
        target_day = today + timedelta(days=1)

    day_month = int(target_day.strftime("%d%m"))
    url = f"https://altask.ru/images/raspisanie/DO/{day_month}.xls"
    response = requests.get(url)
    if response.status_code != 200:
        raise Exception(f"Не удалось скачать файл по ссылке: {url}")
    file_path = f"{day_month}.xls"
    with open(file_path, "wb") as f:
        f.write(response.content)
    parse_and_generate_tables(file_path)

def convert_xls_to_xlsx(input_path):
    try:
        subprocess.run([
            "libreoffice",
            "--headless",
            "--convert-to", "xlsx",
            "--outdir", os.path.dirname("./"),
            input_path
        ], check=True)
        return True
    except subprocess.CalledProcessError as e:
        return False

def read_xls_file(file_path):
    result = []
    try:
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)

        for y in range(sheet.nrows):
            for x in range(sheet.ncols):
                value = sheet.cell_value(y, x)
                if value in GROUP_NAMES:
                    ly = y + 1
                     # изза субботы снизу ебанутой до 40 клеток вниз уходит, пока поставил ограничение 13
                    while ly - y < 13 and ly < sheet.nrows and sheet.cell_value(ly, x) not in GROUP_NAMES:
                        ly += 1
                    if ly >= sheet.nrows:
                        ly = sheet.nrows - 1
                    result.append({
                        "group": value,
                        "x": x,
                        "y1": y + 1,
                        "y2": ly})
        return result
    except Exception as e:
        print(f"[!] Ошибка чтения .xls: {e}")
        return None

from pdf2image import convert_from_path
import subprocess
import os

def create_group_sheets_single_column(groups, source_sheet, output_dir):
    os.makedirs(output_dir, exist_ok=True)

    for group in groups:
        x = group["x"] + 1
        y1 = group["y1"]
        y2 = group["y2"]
        name = group["group"]


        wb = Workbook()
        ws = wb.active
        ws.title = name
        for row in range(y1, y2 + 1):
            src_cell = source_sheet.cell(row=row, column=x)
            tgt_cell = ws.cell(row=row - y1 + 1, column=4, value=src_cell.value)

            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                current_border = src_cell.border
                mixed_border = Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    #TODO
                    top=current_border.top if current_border.top else Side(style="thin", color="505050"),
                    bottom=current_border.bottom if current_border.bottom else Side(style="thin", color="505050")
                )
                tgt_cell.border = mixed_border
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.number_format = copy(src_cell.number_format)
                tgt_cell.protection = copy(src_cell.protection)
                tgt_cell.alignment = copy(src_cell.alignment)

        col_letter_src = source_sheet.cell(row=y1, column=x).column_letter
        ws.column_dimensions['D'].width = max(source_sheet.column_dimensions[col_letter_src].width, 33,22)

        for row in range(y1, y2 + 1):
            if source_sheet.row_dimensions[row].height is not None:
                ws.row_dimensions[row - y1 + 1].height = source_sheet.row_dimensions[row].height

        xlsx_path = os.path.join(output_dir, f"{name}.xlsx")
        wb.save(xlsx_path)
        
        try:
            subprocess.run([
                "libreoffice",
                "--headless",
                "--convert-to", "pdf",
                "--outdir", output_dir,
                xlsx_path
            ], check=True)
            pdf_path = os.path.join(output_dir, f"{name}.pdf")
            
            images = convert_from_path(pdf_path, dpi=300)
            
            png_path = os.path.join(output_dir, f"{name}.png")

            #TODO tut obrezat

            if images:
                img = images[0]
                width, height = img.size
                cropped = img.crop((380, 200, width - 380, height - 400))
                cropped.save(os.path.join(output_dir, f"{name}.png"), "PNG")

            os.remove(xlsx_path)
            os.remove(pdf_path)
            
        except Exception as e:
            print(f"[!] Ошибка при конвертации файлов для группы {name}: {e}")
            continue

    print(f"группы сохранены в {output_dir}")

def parse_and_generate_tables(INPUT_XLS):
    output_dir = "./output"
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    convert_xls_to_xlsx(INPUT_XLS)
    groups = read_xls_file(INPUT_XLS)
    workbook = load_workbook(f"{INPUT_XLS}x")
    sheet = workbook.active
    create_group_sheets_single_column(groups, sheet, output_dir)
    if os.path.exists(INPUT_XLS):
        os.remove(INPUT_XLS)
    if os.path.exists(f"{INPUT_XLS}x"):
        os.remove(f"{INPUT_XLS}x")

    return True